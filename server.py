#!/usr/bin/env python3
"""
投资追踪 Agent - Web 服务器
启动后浏览器打开 http://localhost:5000 即可使用

支持对话 AI（DeepSeek / OpenAI），联网搜索。
"""
import json
import os
import sys
import re
import time
import threading
import urllib.request
import urllib.parse
from datetime import datetime
from http.server import HTTPServer, BaseHTTPRequestHandler
from pathlib import Path

# 确保能找到同目录模块
sys.path.insert(0, str(Path(__file__).parent))
import portfolio_agent as agent

HOST = os.environ.get("HOST", "0.0.0.0")
PORT = int(os.environ.get("PORT", os.environ.get("RAILWAY_PORT", 5050)))
DATA_DIR = Path(__file__).parent
HTML_PATH = DATA_DIR / "webapp.html"

# ============ LLM 配置 ============
# 通过环境变量设置:
#   LLM_PROVIDER=deepseek|openai
#   LLM_API_KEY=sk-xxxxx
#   LLM_MODEL=deepseek-chat|gpt-4o
# 也可以在网页 Agent 的设置页面配置

LLM_CONFIG = {
    "provider": os.environ.get("LLM_PROVIDER", "deepseek"),
    "api_key": os.environ.get("LLM_API_KEY", ""),
    "model": os.environ.get("LLM_MODEL", "deepseek-chat"),
}

LLM_ENDPOINTS = {
    "deepseek": "https://api.deepseek.com/v1/chat/completions",
    "openai": "https://api.openai.com/v1/chat/completions",
}


def _read_config_sheet():
    """从 Excel 配置表读取 LLM 设置"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(agent.EXCEL_PATH)
        if "配置" in wb.sheetnames:
            ws = wb["配置"]
            cfg = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:
                    cfg[str(row[0]).strip()] = str(row[1]).strip()
            wb.close()
            # 如果 Excel 里有 API Key，覆盖环境变量
            if "api_key" in cfg and cfg["api_key"]:
                LLM_CONFIG["api_key"] = cfg["api_key"]
            if "provider" in cfg and cfg["provider"]:
                LLM_CONFIG["provider"] = cfg["provider"]
            if "model" in cfg and cfg["model"]:
                LLM_CONFIG["model"] = cfg["model"]
    except Exception:
        pass


def save_config_sheet(key, value):
    """保存配置到 Excel 配置表"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(agent.EXCEL_PATH)
        if "配置" not in wb.sheetnames:
            wb.create_sheet("配置")
        ws = wb["配置"]
        # 检查是否已有该键
        found = False
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=False):
            if row[0].value and str(row[0].value).strip() == key:
                row[1].value = value
                found = True
                break
        if not found:
            next_row = (ws.max_row or 1) + 1
            ws.cell(row=next_row, column=1, value=key)
            ws.cell(row=next_row, column=2, value=value)
        wb.save(agent.EXCEL_PATH)
        wb.close()
        return True
    except Exception:
        return False


_read_config_sheet()

def _portfolio_summary():
    """生成简短持仓概况字符串"""
    h = agent.load_portfolio()
    tv, tp, accts = agent.calculate(h)
    cash = sum(x["市值"] for x in h if x["类别"] == "现金")
    stocks_str = ", ".join(f"{x['名称']}{x['数量']}股" for x in h if x["类别"] == "股票")
    return f"总资产{tv:,.0f}元 | 现金{cash:,.0f}元 | 累计盈亏{tp:+,.0f}元 | 持仓: {stocks_str}"


def _build_system_prompt():
    """构建包含持仓上下文的 system prompt"""
    from datetime import timezone, timedelta
    now = datetime.now(timezone(timedelta(hours=8)))
    date_str = now.strftime("%Y年%m月%d日 %A")

    holdings = agent.load_portfolio()
    total_value, total_pnl, accounts = agent.calculate(holdings)
    risks = agent.risk_analysis(holdings)

    # 计算股票占大盘子的比例
    def total_alloc(mv):
        return mv / total_value * 100 if total_value > 0 else 0

    # 加载记忆
    memories = agent.load_memories()
    mem_text = ""
    if memories:
        mem_text = "\n".join(f"- {m}" for m in memories[-20:])

    prompt = f"""你是万万的专属投资理财助手，名叫「万万的投资 Agent」。

当前日期：{date_str}（北京时间）

## 你的角色
- 你是专业的个人理财顾问，擅长 A 股投资分析、仓位管理、资产配置
- 回答要专业、简洁、有实操性
- 对风险要明确提示，不做虚假承诺

## 万万当前的持仓

总资产: {total_value:,.0f} 元
累计盈亏: {total_pnl:+,.0f} 元
总收益率: {total_pnl/(total_value-total_pnl)*100:.2f}%"""

    # 股票持仓（仓位占比 = 占总资产比例）
    stock_text = "\n\n### 股票持仓：\n"
    for h in holdings:
        if h["类别"] == "股票" and h["账户"] != "账户二（短线）":
            ret = h["收益率"] * 100 if h["收益率"] else 0
            mv = h["市值"] or 0
            alloc = total_alloc(mv)
            stock_text += f"- {h['名称']}：{h['数量']}股 @ {h['现价']:.2f}元，市值 {mv:,.0f}元，盈亏 {h['盈亏']:+,.0f}元 ({ret:+.1f}%)，占总资产 {alloc:.1f}%\n"
    prompt += stock_text

    # 账户概况
    acc_text = "\n### 各账户：\n"
    for acct, data in sorted(accounts.items()):
        acc_text += f"- {acct}：资产 {data['市值']:,.0f}元，盈亏 {data['盈亏']:+,.0f}元\n"
    prompt += acc_text

    # 资产配置
    by_type = {}
    for h in holdings:
        t = h["类别"]
        by_type[t] = by_type.get(t, 0) + h["市值"]
    prompt += f"\n### 资产配置：\n"
    for t, v in sorted(by_type.items(), key=lambda x: -x[1]):
        prompt += f"- {t}：{v:,.0f}元 ({v/total_value*100:.1f}%)\n"

    from collections import defaultdict
    sector_groups = defaultdict(list)
    for h2 in holdings:
        if h2["类别"] == "股票" and h2["账户"] != "账户二（短线）" and h2["名称"] in agent.SECTOR_MAP:
            sector_groups[agent.SECTOR_MAP[h2["名称"]]].append(h2)
    if sector_groups:
        prompt += "\n### 板块分布：\n"
        for sector, stocks in sorted(sector_groups.items(), key=lambda x: -sum(s["市值"] for s in x[1])):
            names = ", ".join(f"{s['名称']}({s['市值']:,.0f}元)" for s in stocks)
            total_sector = sum(s["市值"] for s in stocks)
            prompt += f"- {sector}：{names}（占总资产 {total_alloc(total_sector):.1f}%）\n"

    # 当前风险提示
    if risks:
        prompt += "\n### 当前诊断的风险：\n"
        for r in risks[:5]:
            prompt += f"- [{r['级别']}] {r['问题']}\n"

    # 近期操作记录
    try:
        edits = agent.load_edit_log()
        if edits:
            prompt += "\n### 近期操作记录：\n"
            for e in edits[-8:]:
                prompt += f"- {e['time']}：{e['detail']}\n"
    except Exception:
        pass

    # 近期出入金
    try:
        flows = agent.load_capital_flows()
        if flows:
            recent = flows[-5:] if len(flows) > 5 else flows
            prompt += "\n### 近期出入金：\n"
            for f in recent:
                sign = "+" if f["type"] == "存入" else "-"
                note = f"（{f['note']}）" if f.get("note") else ""
                prompt += f"- {f['date']} {f['type']} {sign}{f['amount']:,.0f}元{note}\n"
    except Exception:
        pass

    # 净值趋势（最近几天）
    try:
        history = agent.load_history()
        if history:
            recent_h = history[-5:] if len(history) > 5 else history
            prompt += "\n### 近期净值变化：\n"
            for h in recent_h:
                daily = h.get("当日盈亏", 0) or 0
                sign = "+" if daily >= 0 else ""
                prompt += f"- {h['日期']}：总资产 {h['总资产']:,.0f}元，当日盈亏 {sign}{daily:,.0f}元\n"
    except Exception:
        pass

    # AI 记忆
    if mem_text:
        prompt += "\n### 记忆（来自之前的对话）：\n"
        prompt += mem_text + "\n"

    prompt += """

## 重要规则（必须遵守）
1. **A 股最小交易单位是 1手 = 100股**，所有买卖建议的股数必须为 100 的整数倍
2. **仓位百分比以总资产为分母**（不是以股票市值作分母），单票占总资产不超过 25%
3. **给出具体数字**：建议操作时，明确说"卖 X 手（XX 股）"而不是"减仓一部分"
4. **现金为王**：至少保留总资产 10% 的现金
5. **参考操作记录**：回答时可参考用户的近期买卖、出入金操作，分析其操作效果并给出判断

## 回答风格
- 短期建议：具体操作层面（减仓、加仓、止损线）
- 中期建议：配置调整、行业布局
- 长期建议：资产框架、财务规划
- 需要联网信息时，使用 <search>关键词</search> 标明

## 注意
- 你只提供分析建议，不构成投资指令
- 所有操作建议需要结合万万自己的判断
- 数据来自万万手动维护的持仓表，可能有延迟

## 记忆功能
- 你有长期记忆能力，之前的对话要点会显示在"记忆"中
- 如果你想记住用户的重要偏好、决定或结论，在回答末尾加上 <save_memory>内容</save_memory>
- 每次对话开始时你会看到之前保存的记忆"""
    return prompt


def web_search(query, max_results=5):
    """联网搜索，返回搜索结果列表"""
    results = []

    # 用 POST 方式请求 html.duckduckgo.com（比 lite 更稳定）
    for attempt in range(2):
        try:
            if attempt == 0:
                data = urllib.parse.urlencode({"q": query}).encode()
                req = urllib.request.Request(
                    "https://html.duckduckgo.com/html/",
                    data=data,
                    headers={
                        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
                        "Content-Type": "application/x-www-form-urlencoded",
                    }
                )
            else:
                # 备选：lite 模式
                encoded = urllib.parse.quote(query)
                req = urllib.request.Request(
                    f"https://lite.duckduckgo.com/lite/?q={encoded}",
                    headers={"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"}
                )

            resp = urllib.request.urlopen(req, timeout=15)
            html = resp.read().decode("utf-8", errors="ignore")

            if attempt == 0:
                # 解析 HTML 模式：找 result__a（标题链接）和 result__snippet
                titles_links = re.findall(r'<a[^>]*class="result__a"[^>]*href="([^"]*)"[^>]*>(.*?)</a>', html, re.DOTALL)
                snippets = re.findall(r'<a[^>]*class="result__snippet"[^>]*>(.*?)</a>', html, re.DOTALL)
                for i in range(min(len(titles_links), max_results)):
                    url = titles_links[i][0]
                    title = re.sub(r'<[^>]+>', '', titles_links[i][1]).strip()
                    snippet = re.sub(r'<[^>]+>', '', snippets[i]).strip() if i < len(snippets) else ""
                    if url and not url.startswith("http"):
                        url = "https://" + url
                    results.append({"title": title, "url": url, "snippet": snippet})

                # 如果上面没解析到，试试更宽松的匹配
                if not results:
                    all_links = re.findall(r'<a[^>]*class="result[^"]*"[^>]*href="([^"]*)"[^>]*>(.*?)</a>', html, re.DOTALL)
                    for i in range(min(len(all_links), max_results)):
                        url = all_links[i][0]
                        title = re.sub(r'<[^>]+>', '', all_links[i][1]).strip()
                        if url and not url.startswith("http"):
                            url = "https://" + url
                        results.append({"title": title, "url": url, "snippet": ""})
            else:
                # 解析 Lite 模式
                links = re.findall(r'<a[^>]*href="([^"]*)"[^>]*class="result-link"[^>]*>(.*?)</a>', html, re.DOTALL)
                snippets = re.findall(r'<td[^>]*class="result-snippet"[^>]*>(.*?)</td>', html, re.DOTALL)
                for i in range(min(len(links), max_results)):
                    url = links[i][0]
                    title = re.sub(r'<[^>]+>', '', links[i][1]).strip()
                    snippet = re.sub(r'<[^>]+>', '', snippets[i]).strip() if i < len(snippets) else ""
                    if url and not url.startswith("http"):
                        url = "https://" + url
                    results.append({"title": title, "url": url, "snippet": snippet})

            if results:
                break
        except Exception as e:
            print(f"⚠️ 搜索尝试 {attempt+1} 失败: {e}")
            continue

    return results


def chat_stream(messages, search_enabled=True):
    """流式调用 LLM API，逐 block 产出文本"""
    api_key = LLM_CONFIG["api_key"]
    if not api_key:
        yield json.dumps({"error": "未配置 API Key，请在设置页面填写"}, ensure_ascii=False)
        return

    provider = LLM_CONFIG["provider"]
    model = LLM_CONFIG["model"]
    endpoint = LLM_ENDPOINTS.get(provider, LLM_ENDPOINTS["deepseek"])
    if not endpoint:
        yield json.dumps({"error": f"未知 provider: {provider}"}, ensure_ascii=False)
        return

    # 构建消息列表
    system_prompt = _build_system_prompt()
    full_messages = [{"role": "system", "content": system_prompt}]

    # 检查最后一条消息是否需要联网搜索
    last_msg = messages[-1]["content"] if messages else ""
    search_keywords = ["搜索", "查一下", "联网", "最新", "今天", "行情", "新闻", "股票", "基金", "涨", "跌"]
    should_search = search_enabled and any(kw in last_msg for kw in search_keywords)

    search_results = None
    if should_search:
        yield json.dumps({"type": "search_status", "content": "🔍 正在联网搜索..."}, ensure_ascii=False)

        # 异步搜索：后台线程 + 心跳保活，防止 Railway 超时断开 SSE
        result_box = []
        def _do_search():
            result_box.append(web_search(last_msg))
        t = threading.Thread(target=_do_search, daemon=True)
        t.start()

        phases = ["⏳ 搜索中", "⏳ 搜索中.", "⏳ 搜索中..", "⏳ 搜索中..."]
        idx = 0
        max_wait = 35
        start = time.time()
        while t.is_alive():
            if time.time() - start > max_wait:
                break
            t.join(timeout=2.5)
            idx = (idx + 1) % len(phases)
            yield json.dumps({"type": "search_heartbeat", "content": phases[idx]}, ensure_ascii=False)

        if result_box:
            search_results = result_box[0]

        if search_results:
            context = "\n\n### 联网搜索结果：\n"
            for i, r in enumerate(search_results, 1):
                context += f"{i}. [{r['title']}]({r['url']})\n   {r['snippet']}\n"
            full_messages[0]["content"] += context
            yield json.dumps({"type": "search_status", "content": f"✅ 搜索到 {len(search_results)} 条结果，正在分析..."}, ensure_ascii=False)
        else:
            yield json.dumps({"type": "search_status", "content": "⚠️ 搜索结果为空，直接分析"}, ensure_ascii=False)

    full_messages.extend(messages)

    # 准备请求
    req_body = json.dumps({
        "model": model,
        "messages": full_messages,
        "stream": True,
        "max_tokens": 4096,
        "temperature": 0.7,
    }).encode("utf-8")

    req = urllib.request.Request(
        endpoint,
        data=req_body,
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}",
            "Accept": "text/event-stream",
        },
        method="POST",
    )

    try:
        resp = urllib.request.urlopen(req, timeout=60)
        buffer = ""
        while True:
            chunk = resp.read(1024)
            if not chunk:
                break
            buffer += chunk.decode("utf-8", errors="ignore")
            while "\n" in buffer:
                line, buffer = buffer.split("\n", 1)
                line = line.strip()
                if not line or line.startswith(":"):
                    continue
                if line == "data: [DONE]":
                    return
                if line.startswith("data: "):
                    try:
                        data = json.loads(line[6:])
                        delta = data.get("choices", [{}])[0].get("delta", {})
                        content = delta.get("content", "")
                        if content:
                            yield json.dumps({"type": "text", "content": content}, ensure_ascii=False)
                    except json.JSONDecodeError:
                        continue
    except urllib.error.HTTPError as e:
        err_body = e.read().decode("utf-8", errors="ignore")
        yield json.dumps({"type": "error", "content": f"API 请求失败 (HTTP {e.code}): {err_body[:200]}"}, ensure_ascii=False)
    except Exception as e:
        yield json.dumps({"type": "error", "content": f"请求异常: {str(e)}"}, ensure_ascii=False)


# ============ API 处理 ============

def _fetch_prev_closes():
    """从腾讯 API 获取所有股票的昨收价"""
    import urllib.request
    holdings = agent.load_portfolio()
    codes = {}
    for h in holdings:
        if h["类别"] == "股票" and h.get("代码"):
            codes[h["名称"]] = h["代码"]
    for name, code in agent.STOCK_CODES.items():
        if name not in codes:
            codes[name] = code
    if not codes:
        return {}
    tdx_codes = []
    code_to_name = {}
    for name, code in codes.items():
        for variant in agent._code_variants(code):
            tdx_codes.append(variant)
            code_to_name[variant] = name
    url = "http://qt.gtimg.cn/q=" + ",".join(tdx_codes)
    prev_closes = {}
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        resp = urllib.request.urlopen(req, timeout=8)
        text = resp.read().decode("gbk")
        for line in text.strip().split(";"):
            line = line.strip()
            if not line or "=" not in line:
                continue
            parts = line.split("~")
            if len(parts) >= 5:
                ticker = parts[0].split("=")[0].strip("v_")
                name_from_api = parts[1] if len(parts) > 1 else ""
                if ticker in code_to_name and name_from_api and name_from_api != "?":
                    name = code_to_name[ticker]
                    try:
                        pc = float(parts[4]) if parts[4] else 0
                        if pc > 0 and name not in prev_closes:
                            prev_closes[name] = pc
                    except ValueError:
                        pass
    except Exception:
        pass
    return prev_closes


def api_data():
    """返回持仓数据 JSON"""
    holdings = agent.load_portfolio()
    total_value, total_pnl, accounts = agent.calculate(holdings)
    history = agent.load_history()
    prev_closes = _fetch_prev_closes()
    yesterday_qty = agent.load_yesterday_qty_map()  # 名称 → 昨日持股数

    stocks = []
    for h in holdings:
        if h["类别"] == "股票":
            qty = h["数量"] or 0
            if qty == 0:
                continue  # 已清仓的不显示在持仓中
            price = h["现价"] or 0
            prev_close = prev_closes.get(h["名称"], 0)
            # 今日新增的股票（昨日持股数=0），用成本价代替昨收价计算今日收益
            yesterday_q = yesterday_qty.get(h["名称"], None)
            is_new_today = yesterday_q is not None and yesterday_q == 0
            if is_new_today:
                today_pnl = round((price - h["成本价"]) * qty, 2) if qty > 0 else 0
            else:
                today_pnl = round((price - prev_close) * qty, 2) if prev_close > 0 and qty > 0 else 0
            stocks.append({
                "name": h["名称"],
                "account": h["账户"],
                "price": price,
                "cost": h["成本价"],
                "qty": qty,
                "market_value": h["市值"],
                "pnl": h["盈亏"],
                "return_pct": round(h["收益率"] * 100, 2) if h["收益率"] else 0,
                "alloc_pct": round(h["仓位占比"] * 100, 1) if h["仓位占比"] else 0,
                "today_pnl": today_pnl,
            })

    by_type = {}
    for h in holdings:
        t = h["类别"]
        by_type[t] = by_type.get(t, 0) + h["市值"]

    # 完整原始数据（包含黄金、现金等，供编辑弹窗使用）
    raw_holdings = []
    for h in holdings:
        raw_holdings.append({
            "type": h["类别"],
            "account": h["账户"],
            "name": h["名称"],
            "code": h.get("代码", ""),
            "qty": h["数量"],
            "cost": h["成本价"],
            "market_value": h.get("市值", 0),
            "note": h.get("备注", ""),
        })

    net_capital, flow_count = agent.get_net_capital()
    true_pnl = round(total_value - net_capital, 2) if net_capital > 0 else total_pnl

    # 当日出入金净额（用于调整今日盈亏）
    today = datetime.now().strftime("%Y-%m-%d")
    all_flows = agent.load_capital_flows()
    today_net_flow = sum(
        f["amount"] if f["type"] == "存入" else -f["amount"]
        for f in all_flows if f["date"].startswith(today)
    )

    realized_pnl = agent.get_realized_pnl_total()
    unrealized_pnl = total_pnl - realized_pnl

    # 已清仓股票（持股数=0）
    cleared_stocks = agent.get_cleared_positions()

    return {
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "total_value": total_value,
        "total_pnl": total_pnl,
        "unrealized_pnl": unrealized_pnl,
        "realized_pnl": realized_pnl,
        "total_return_pct": round(total_pnl / (total_value - total_pnl) * 100, 2) if (total_value - total_pnl) > 0 else 0,
        "net_capital": net_capital,
        "true_pnl": true_pnl,
        "today_net_flow": today_net_flow,
        "stocks": stocks,
        "cleared_stocks": cleared_stocks,
        "accounts": {k: {"value": v["市值"], "pnl": v["盈亏"], "stock_value": v.get("股票市值", 0)} for k, v in accounts.items()},
        "allocation": {k: v for k, v in sorted(by_type.items(), key=lambda x: -x[1])},
        "history": [{"date": r["日期"], "weekday": agent._weekday_cn(r["日期"]), "total": r["总资产"], "pnl": r["累计盈亏"], "daily": r["当日盈亏"]} for r in history],
        "holdings_raw": raw_holdings,
    }


def api_refresh():
    """刷新行情并返回更新后的数据"""
    holdings = agent.load_portfolio()
    prices, prev_closes, failed = agent.fetch_prices()
    if prices:
        holdings = agent.update_holdings_with_prices(holdings, prices, prev_closes)
        agent.save_prices(holdings)
    # 记录净值（只在15:00收盘后记录）
    total_value, total_pnl, accounts = agent.calculate(holdings)
    detail = {k: v["市值"] for k, v in accounts.items()}
    snapshot_recorded = agent.save_networth_snapshot(total_value, total_pnl, detail)

    # 每次刷新都同步 Excel 到 GitHub，防止部署丢数据
    try:
        agent.sync_portfolio_to_github()
    except Exception:
        pass
    # 更新 HTML 看板
    try:
        from report_generator import generate as gen_html
        gen_html()
    except Exception:
        pass
    return {
        "success": True,
        "updated": len(prices),
        "failed": failed,
        "snapshot_recorded": snapshot_recorded,
        "data": api_data(),
    }


def api_advice():
    """返回投资建议"""
    holdings = agent.load_portfolio()
    risks = agent.risk_analysis(holdings)
    advice_text = agent.generate_advice(holdings, risks)
    agent.save_advice(advice_text)
    risks_out = [{"level": r["级别"], "issue": r["问题"], "suggestion": r["建议"]} for r in risks]
    return {
        "risks": risks_out,
        "text": advice_text,
    }


def api_breakeven():
    """返回回本分析"""
    holdings = agent.load_portfolio()
    cash = sum(h["市值"] for h in holdings if h["类别"] == "现金")
    return {"stocks": agent.breakeven_analysis(holdings, cash)}


def api_kline():
    """返回持仓股票的日K线数据"""
    try:
        return {"data": agent.fetch_kline()}
    except Exception as e:
        return {"error": str(e), "data": {}}


def api_breakeven_realistic(data):
    """AI分析某只股票基于市场行情的真实回本预估"""
    name = data.get("name", "")
    sector = data.get("sector", "")
    cost = data.get("cost", 0)
    price = data.get("price", 0)
    rise_needed = (cost / price - 1) * 100 if price > 0 else 0
    if not name or not sector:
        return {"error": "参数不全"}

    api_key = LLM_CONFIG["api_key"]
    if not api_key:
        return {"scenarios": None, "error": "未配置 API Key"}

    provider = LLM_CONFIG["provider"]
    model = LLM_CONFIG["model"]
    endpoint = LLM_ENDPOINTS.get(provider, LLM_ENDPOINTS["deepseek"])

    # 搜索市场消息
    queries = [f"{name} 股票 走势 分析 {datetime.now().strftime('%Y-%m')}", f"{sector} 板块 行情 展望"]
    news = []
    for q in queries:
        for r in web_search(q, 4):
            news.append(f"- [{r['title']}] {r['snippet']}")
    context = "\n".join(news[:8]) or "无搜索结果"

    prompt = f"""你是A股技术分析师。分析以下股票的合理回本预期。

股票：{name}
板块：{sector}
成本价：{cost}元
现价：{price}元（需涨 {rise_needed:.1f}% 回本）

## 最新市场信息
{context}

## 请基于以上市场信息，输出三个场景的合理月涨幅预估（纯JSON）：
{{{{
  "optimistic": {{"monthly_rise": 数字, "reason": "简析"}},
  "moderate": {{"monthly_rise": 数字, "reason": "简析"}},
  "conservative": {{"monthly_rise": 数字, "reason": "简析"}}
}}}}

要求：
- 每月涨幅基于真实市场情况（板块热度、个股消息等）
- optimistic 3-8%，moderate 1-4%，conservative 0.3-1.5%
- reason 用一句话说明判断依据，必须用中文
- 输出纯JSON，不要其他文字"""

    messages = [
        {"role": "system", "content": "你必须用中文回复，输出纯JSON，不要任何其他文字。"},
        {"role": "user", "content": prompt},
    ]

    try:
        req_body = json.dumps({"model": model, "messages": messages, "stream": False, "max_tokens": 1024, "temperature": 0.3}).encode()
        req = urllib.request.Request(endpoint, data=req_body, headers={"Content-Type": "application/json", "Authorization": f"Bearer {api_key}"}, method="POST")
        resp = urllib.request.urlopen(req, timeout=60)
        result = json.loads(resp.read().decode("utf-8"))
        content = result["choices"][0]["message"]["content"]
        m = re.search(r'\{.*\}', content, re.DOTALL)
        scenarios = json.loads(m.group()) if m else json.loads(content)

        # 计算回本天数
        out = {}
        for k, v in scenarios.items():
            mr = v["monthly_rise"]
            if mr > 0 and rise_needed > 0:
                days = round(rise_needed / (mr / 22))
            else:
                days = 999
            label_map = {"optimistic":"乐观","moderate":"中性","conservative":"保守"}
            out[k] = {"label": label_map.get(k, k), "days": days, "monthly_rise": mr, "reason": v.get("reason", "")}
        return {"scenarios": out, "rise_needed": round(rise_needed, 1), "name": name}
    except Exception as e:
        return {"error": str(e), "scenarios": None}


def api_add_position_advice(data):
    """AI 分析某只股票是否应该加仓，基于市场消息和行情"""
    stock_name = data.get("name", "")
    sector = data.get("sector", "")
    qty = data.get("qty", 0)
    cost_price = data.get("cost", 0)
    current_price = data.get("price", 0)
    cash = data.get("cash", 0)

    api_key = LLM_CONFIG["api_key"]
    if not api_key:
        return {"error": "未配置 API Key，请在设置页面填写"}

    provider = LLM_CONFIG["provider"]
    model = LLM_CONFIG["model"]
    endpoint = LLM_ENDPOINTS.get(provider, LLM_ENDPOINTS["deepseek"])

    # 搜索股票 + 板块消息
    search_queries = [
        f"{stock_name} 股票 最新消息 {datetime.now().strftime('%Y-%m')}",
        f"{sector} 板块 最新行情 分析",
    ]
    news_items = []
    for q in search_queries:
        results = web_search(q, max_results=4)
        for r in results:
            news_items.append(f"- [{r['title']}] {r['snippet']}")

    search_context = "\n".join(news_items[:8]) if news_items else "未搜索到相关消息"

    rise_needed = (cost_price / current_price - 1) * 100 if current_price > 0 else 0

    portfolio_summary = _portfolio_summary()

    prompt = f"""你是专业的A股投资顾问。请分析以下持仓是否应该加仓。

## 万万整体持仓
{portfolio_summary}

## 持仓信息
- 股票：{stock_name}
- 所属板块：{sector}
- 持有数量：{qty} 股
- 成本均价：{cost_price} 元
- 当前价格：{current_price} 元
- 需涨 {rise_needed:.1f}% 才能回本
- 可用现金：{cash:,.0f} 元

## 最新市场消息
{search_context}

## 任务
基于以上持仓数据和最新的市场消息、板块行情，给出加仓建议。
以 JSON 格式输出（不要其他文字）：
{{{{
  "should_add": true/false,
  "timing": "立即加仓/分批次加仓/观望等回调/不建议加仓",
  "suggested_lots": 建议几手（0表示不加，1手=100股），
  "reason": "判断理由（结合市场消息具体说明）",
  "risk_tip": "风险提示"
}}}}"""

    messages = [
        {"role": "system", "content": "你是严谨的A股投资顾问，总是基于最新市场信息做判断。必须用中文，输出纯 JSON，不要任何其他文字。"},
        {"role": "user", "content": prompt},
    ]

    req_body = json.dumps({
        "model": model,
        "messages": messages,
        "stream": False,
        "max_tokens": 1024,
        "temperature": 0.3,
    }).encode("utf-8")

    try:
        req = urllib.request.Request(
            endpoint, data=req_body,
            headers={"Content-Type": "application/json",
                     "Authorization": f"Bearer {api_key}"},
            method="POST",
        )
        resp = urllib.request.urlopen(req, timeout=60)
        result = json.loads(resp.read().decode("utf-8"))
        content = result["choices"][0]["message"]["content"]

        # 提取 JSON
        json_match = re.search(r'\{.*\}', content, re.DOTALL)
        advice = json.loads(json_match.group()) if json_match else json.loads(content)
        advice["stock_name"] = stock_name
        advice["sector"] = sector
        return advice
    except Exception as e:
        return {"error": f"AI 分析失败: {str(e)}", "stock_name": stock_name}


# ============ HTTP 服务器 ============

class PortfolioHandler(BaseHTTPRequestHandler):
    def _send_json(self, data, status=200):
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _send_html(self, html):
        body = html.encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _read_html(self):
        if not HTML_PATH.exists():
            return "<html><body><h2>webapp.html 未找到</h2></body></html>"
        return HTML_PATH.read_text(encoding="utf-8")

    def do_GET(self):
        if self.path == "/" or self.path == "/index.html":
            self._send_html(self._read_html())

        elif self.path == "/api/data":
            try:
                self._send_json(api_data())
            except Exception as e:
                self._send_json({"error": str(e)}, 500)

        elif self.path == "/api/refresh":
            try:
                self._send_json(api_refresh())
            except Exception as e:
                self._send_json({"error": str(e)}, 500)

        elif self.path == "/api/advice":
            try:
                self._send_json(api_advice())
            except Exception as e:
                self._send_json({"error": str(e)}, 500)

        elif self.path == "/api/breakeven":
            try:
                self._send_json(api_breakeven())
            except Exception as e:
                self._send_json({"error": str(e)}, 500)

        elif self.path == "/api/kline":
            try:
                self._send_json(api_kline())
            except Exception as e:
                self._send_json({"error": str(e)}, 500)

        elif self.path == "/api/capital-flows":
            try:
                flows = agent.load_capital_flows()
                net, cnt = agent.get_net_capital()
                self._send_json({"flows": flows, "net_capital": net})
            except Exception as e:
                self._send_json({"error": str(e)}, 500)

        elif self.path == "/api/config":
            self._send_json({
                "provider": LLM_CONFIG["provider"],
                "model": LLM_CONFIG["model"],
                "has_key": bool(LLM_CONFIG["api_key"]),
                "key_prefix": LLM_CONFIG["api_key"][:8] + "..." if LLM_CONFIG["api_key"] else "",
            })

        elif self.path == "/api/realized-pnl":
            try:
                records = agent.get_realized_pnl_records()
                total = agent.get_realized_pnl_total()
                self._send_json({"records": records, "total": total})
            except Exception as e:
                self._send_json({"error": str(e)}, 500)

        elif self.path == "/api/edit-history":
            try:
                logs = agent.load_edit_log()
                self._send_json({"logs": logs})
            except Exception as e:
                self._send_json({"error": str(e)}, 500)

        else:
            self._send_json({"error": "not found"}, 404)

    def do_POST(self):
        content_length = int(self.headers.get("Content-Length", 0))
        body = self.rfile.read(content_length) if content_length > 0 else b"{}"
        data = json.loads(body) if body else {}

        if self.path == "/api/chat":
            self.send_response(200)
            self.send_header("Content-Type", "text/event-stream; charset=utf-8")
            self.send_header("Cache-Control", "no-cache")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()

            messages = data.get("messages", [])
            search_enabled = data.get("search", True)
            full_response = ""

            for chunk in chat_stream(messages, search_enabled):
                try:
                    # 累积文本，提取记忆
                    try:
                        cdata = json.loads(chunk) if chunk.startswith("{") else {}
                        if cdata.get("type") == "text":
                            full_response += cdata.get("content", "")
                    except Exception:
                        pass
                    self.wfile.write(f"data: {chunk}\n\n".encode("utf-8"))
                    self.wfile.flush()
                except BrokenPipeError:
                    break
            try:
                self.wfile.write(b"data: [DONE]\n\n")
                self.wfile.flush()
            except BrokenPipeError:
                pass

            # 保存记忆
            mem_match = re.findall(r'<save_memory>(.*?)</save_memory>', full_response, re.DOTALL)
            for mem_content in mem_match:
                mem_content = mem_content.strip()
                if mem_content:
                    try:
                        agent.save_memory(mem_content)
                        print(f"🧠 保存记忆: {mem_content[:50]}...")
                    except Exception:
                        pass

        elif self.path == "/api/save-config":
            key = data.get("key", "")
            value = data.get("value", "")
            if key == "api_key":
                LLM_CONFIG["api_key"] = value
                save_config_sheet("api_key", value)
            elif key == "provider":
                LLM_CONFIG["provider"] = value
                save_config_sheet("provider", value)
            elif key == "model":
                LLM_CONFIG["model"] = value
                save_config_sheet("model", value)
            self._send_json({"success": True})

        elif self.path == "/api/capital-flow":
            try:
                amount = float(data.get("amount", 0))
                flow_type = data.get("type", "存入")
                note = data.get("note", "")
                if amount <= 0:
                    self._send_json({"success": False, "error": "金额必须大于0"})
                    return
                ok = agent.add_capital_flow(amount, flow_type, note)
                agent.sync_portfolio_to_github()
                self._send_json({"success": ok})
            except Exception as e:
                self._send_json({"success": False, "error": str(e)}, 500)

        elif self.path == "/api/capital-flows":
            try:
                flows = agent.load_capital_flows()
                net, cnt = agent.get_net_capital()
                self._send_json({"flows": flows, "net_capital": net})
            except Exception as e:
                self._send_json({"error": str(e)}, 500)

        elif self.path == "/api/add-position-advice":
            try:
                result = api_add_position_advice(data)
                self._send_json(result)
            except Exception as e:
                self._send_json({"error": str(e)}, 500)

        elif self.path == "/api/holdings-save":
            try:
                holdings_data = data.get("holdings", [])
                result = agent.save_holdings(holdings_data)
                prices, prev_closes, failed = agent.fetch_prices()
                if prices:
                    result = agent.update_holdings_with_prices(result, prices, prev_closes)
                    agent.save_prices(result)
                # 同步到 GitHub，防止 Railway 部署时数据丢失
                agent.sync_portfolio_to_github()
                self._send_json({"success": True})
            except Exception as e:
                self._send_json({"success": False, "error": str(e)}, 500)

        elif self.path == "/api/breakeven-realistic":
            try:
                result = api_breakeven_realistic(data)
                self._send_json(result)
            except Exception as e:
                self._send_json({"error": str(e)}, 500)

        elif self.path == "/api/realized-pnl":
            try:
                action = data.get("action", "create")
                if action == "update":
                    row = data.get("row")
                    name = data.get("name", "").strip()
                    amount = data.get("amount")
                    if not row:
                        self._send_json({"success": False, "error": "缺少行号"})
                        return
                    ok = agent.update_realized_pnl(int(row), name=name or None, amount=amount)
                    if ok:
                        agent.sync_portfolio_to_github()
                    self._send_json({"success": ok})
                elif action == "delete":
                    row = data.get("row")
                    if not row:
                        self._send_json({"success": False, "error": "缺少行号"})
                        return
                    ok = agent.delete_realized_pnl(int(row))
                    if ok:
                        agent.sync_portfolio_to_github()
                    self._send_json({"success": ok})
                else:
                    name = data.get("name", "").strip()
                    amount = float(data.get("amount", 0))
                    if not name or amount == 0:
                        self._send_json({"success": False, "error": "请选择股票并输入盈亏金额"})
                        return
                    agent.record_realized_pnl(name, amount=amount)
                    agent.sync_portfolio_to_github()
                    self._send_json({"success": True})
            except Exception as e:
                self._send_json({"success": False, "error": str(e)}, 500)

        else:
            self._send_json({"error": "not found"}, 404)

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type, Authorization")
        self.end_headers()

    def log_message(self, format, *args):
        print(f"[{datetime.now().strftime('%H:%M:%S')}] {args[0]} {args[1]} {args[2]}")


def main():
    HTTPServer.allow_reuse_address = True
    server = HTTPServer((HOST, PORT), PortfolioHandler)
    print(f"\n{'='*50}")
    print(f"  📊 投资追踪 Agent 已启动！")
    print(f"  {'='*50}")
    print(f"  浏览器打开:")
    print(f"  → http://{HOST}:{PORT}")
    print(f"  {'='*50}")
    print(f"  按 Ctrl+C 停止服务器\n")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n🛑 服务器已停止")
        server.server_close()


def _pull_latest_from_github():
    """启动时从 GitHub 拉取最新的 portfolio.xlsx，防止部署导致数据丢失"""
    import urllib.request
    token = os.environ.get("GH_TOKEN")
    if not token:
        return
    repo = "wanchangjia-png/investment-agent"
    url = f"https://api.github.com/repos/{repo}/contents/portfolio.xlsx"
    try:
        req = urllib.request.Request(url, headers={
            "Authorization": f"Bearer {token}",
            "User-Agent": "investment-agent",
            "Accept": "application/vnd.github.v3+json",
        })
        resp = urllib.request.urlopen(req, timeout=15)
        import json, base64
        data = json.loads(resp.read().decode("utf-8"))
        content = base64.b64decode(data["content"])
        with open(agent.EXCEL_PATH, "wb") as f:
            f.write(content)
        print("📥 已从 GitHub 拉取最新 portfolio.xlsx")
    except Exception as e:
        print(f"⚠️ 从 GitHub 拉取失败（首次部署或 token 问题）: {e}")


if __name__ == "__main__":
    _pull_latest_from_github()
    main()
