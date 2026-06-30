#!/usr/bin/env python3
"""
投资追踪 Agent — 读取 portfolio.xlsx，拉取市场数据，计算盈亏，生成报告。
支持独立运行（CLI）和被 Claude 调用。

用法:
  python3 portfolio_agent.py holdings     # 查看持仓
  python3 portfolio_agent.py refresh      # 更新实时价格
  python3 portfolio_agent.py report       # 生成完整报告
  python3 portfolio_agent.py allocation   # 资产配置分析
  python3 portfolio_agent.py history      # 净值历史趋势
  python3 portfolio_agent.py snapshot     # 保存当日净值快照
"""
import sys
import os
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

# ============ 配置 ============
DATA_DIR = Path(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = DATA_DIR / "portfolio.xlsx"

# A股代码映射（名称 → 代码）
STOCK_CODES = {
    "科达利": "002850",
    "三花智控": "002050",
    "拓普集团": "601689",
    "立讯精密": "002475",
    "五粮液": "000858",
    "天齐锂业": "002466",
    "元隆雅图": "002878",
}

# 板块映射（名称 → 板块）
SECTOR_MAP = {
    "科达利": "新能源车（锂电池结构件）",
    "三花智控": "新能源车（热管理）",
    "拓普集团": "新能源车（底盘系统）",
    "天齐锂业": "新能源车（上游锂矿）",
    "立讯精密": "消费电子/科技",
    "五粮液": "白酒/消费",
    "元隆雅图": "文化传媒/促销品",
}

SECTOR_OUTLOOK = {
    "新能源车（锂电池结构件）": "行业竞争加剧、产能结构性过剩，但龙头份额持续集中；关注下游需求增速和价格战对利润的影响。",
    "新能源车（热管理）": "受益于新能源车渗透率提升和热泵技术升级，长期需求确定；短期受整车降价传导压力。",
    "新能源车（底盘系统）": "线控底盘是智能化核心赛道，ASP持续提升；但需关注客户集中度和订单执行节奏。",
    "新能源车（上游锂矿）": "锂价经历大幅回落，短期供需仍偏宽松；中期看锂资源战略价值，关注成本支撑和需求复苏信号。",
    "消费电子/科技": "AI端侧（手机/PC）驱动换机周期，果链受益确定性较高；但需关注地缘政治和供应链外迁风险。",
    "白酒/消费": "行业处于去库存周期后期，商务消费复苏偏弱；高端白酒抗风险能力强，关注经济复苏节奏和动销数据。",
    "文化传媒/促销品": "行业与宏观经济和消费信心相关，竞争格局分散；关注大客户订单和数字化转型进展。",
}

# 板块预期（通用宏观判断，仅供参考）
SECTOR_OUTLOOK = {
    "新能源车（锂电池结构件）": "行业竞争加剧、产能结构性过剩，但龙头份额持续集中；关注下游需求增速和价格战对利润的影响。",
    "新能源车（热管理）": "受益于新能源车渗透率提升和热泵技术升级，长期需求确定；短期受整车降价传导压力。",
    "新能源车（底盘系统）": "线控底盘是智能化核心赛道，ASP持续提升；但需关注客户集中度和订单执行节奏。",
    "新能源车（上游锂矿）": "锂价经历大幅回落，短期供需仍偏宽松；中期看锂资源战略价值，关注成本支撑和需求复苏信号。",
    "消费电子/科技": "AI端侧（手机/PC）驱动换机周期，果链受益确定性较高；但需关注地缘政治和供应链外迁风险。",
    "白酒/消费": "行业处于去库存周期后期，商务消费复苏偏弱；高端白酒抗风险能力强，关注经济复苏节奏和动销数据。",
}


# ============ Excel 读写 ============
def load_portfolio():
    """加载持仓表，返回持仓列表和合计信息"""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["持仓表"]
    holdings = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or (isinstance(row[0], str) and row[0] == "合计"):
            break
        holdings.append({
            "类别": row[0] or "",
            "账户": row[1] or "",
            "名称": row[2] or "",
            "代码": row[3] or "",
            "数量": row[4] if row[4] != "" else 0,
            "成本价": row[5] if row[5] != "" else 0,
            "现价": row[6] if row[6] != "" else 0,
            "市值": row[7] if row[7] != "" else 0,
            "盈亏": row[8] if row[8] != "" else 0,
            "收益率": row[9] if row[9] != "" else 0,
            "仓位占比": row[10] if row[10] != "" else 0,
            "备注": row[11] or "",
        })
    wb.close()
    return holdings


def save_holdings(holdings_data):
    """保存持仓数据到 Excel（用户手动编辑用）
    holdings_data: list of dicts with keys: 类别, 账户, 名称, 代码, 数量, 成本价, 现价, 备注
    """
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws_old = wb["持仓表"]
        price_map = {}
        for row in ws_old.iter_rows(min_row=2, values_only=True):
            if row[0] == "股票" and row[2] and row[6]:
                price_map[str(row[2]).strip()] = row[6]
        wb.close()
    except Exception:
        price_map = {}

    new_holdings = []
    for d in holdings_data:
        name = d["名称"]
        qty = d.get("数量") or 0
        cost = d.get("成本价") or 0
        price = d.get("现价") or price_map.get(name, cost)

        if d["类别"] == "股票":
            if qty > 0 and cost > 0:
                market_value = round(qty * price, 2)
                cost_total = round(qty * cost, 2)
                pnl = round(market_value - cost_total, 2)
                return_pct = round(pnl / cost_total, 4) if cost_total > 0 else 0
            else:
                market_value = 0; pnl = 0; return_pct = 0
        elif d["类别"] == "现金":
            market_value = d.get("市值") or (qty if qty else 0)
            pnl = 0; return_pct = 0; cost = None
        else:
            market_value = qty * price if qty and price else 0
            pnl = 0; return_pct = 0

        new_holdings.append({
            "类别": d["类别"], "账户": d["账户"], "名称": name,
            "代码": d.get("代码", ""), "数量": qty, "成本价": cost,
            "现价": price, "市值": market_value, "盈亏": pnl,
            "收益率": return_pct, "仓位占比": 0, "备注": d.get("备注", ""),
        })

    total_stock_value = sum(h["市值"] for h in new_holdings
        if h["类别"] == "股票" and h["账户"] != "账户二（短线）")
    for h in new_holdings:
        if h["类别"] == "股票" and h["账户"] != "账户二（短线）" and total_stock_value > 0:
            h["仓位占比"] = round(h["市值"] / total_stock_value, 4)
        elif h["类别"] == "现金":
            h["仓位占比"] = 0

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["持仓表"]
    while ws.max_row > 1:
        ws.delete_rows(2)
    for i, h in enumerate(new_holdings, 2):
        ws.cell(row=i, column=1, value=h["类别"])
        ws.cell(row=i, column=2, value=h["账户"])
        ws.cell(row=i, column=3, value=h["名称"])
        ws.cell(row=i, column=4, value=h["代码"])
        ws.cell(row=i, column=5, value=h["数量"] if h["数量"] else None)
        ws.cell(row=i, column=6, value=h["成本价"] if h["成本价"] else None)
        ws.cell(row=i, column=7, value=h["现价"] if h["现价"] else None)
        ws.cell(row=i, column=8, value=h["市值"] if h["市值"] else None)
        ws.cell(row=i, column=9, value=h["盈亏"] if h["盈亏"] else None)
        ws.cell(row=i, column=10, value=h["收益率"] if h["收益率"] else None)
        ws.cell(row=i, column=11, value=h["仓位占比"] if h["仓位占比"] else None)
        ws.cell(row=i, column=12, value=h["备注"] if h["备注"] else None)
    tr = len(new_holdings) + 2
    ws.cell(row=tr, column=1, value="合计")
    ws.cell(row=tr, column=8, value=sum(h["市值"] for h in new_holdings))
    ws.cell(row=tr, column=9, value=sum(h["盈亏"] for h in new_holdings))
    wb.save(EXCEL_PATH)
    wb.close()
    return load_portfolio()


def save_prices(holdings):
    """将更新后的价格写回 Excel"""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["持仓表"]
    for i, h in enumerate(holdings, 2):
        ws.cell(row=i, column=5, value=h.get("数量"))
        ws.cell(row=i, column=6, value=h.get("成本价"))
        ws.cell(row=i, column=7, value=h["现价"])
        ws.cell(row=i, column=8, value=h["市值"])
        ws.cell(row=i, column=9, value=h["盈亏"])
        ws.cell(row=i, column=10, value=h["收益率"])
        ws.cell(row=i, column=11, value=h["仓位占比"])
    wb.save(EXCEL_PATH)
    wb.close()


def save_networth_snapshot(total, pnl, detail):
    """保存净值快照到净值历史表（只在15:00收盘后记录，一天一次）
    返回 True=已记录, False=跳过"""
    now = datetime.now()
    market_closed = now.hour > 15 or (now.hour == 15 and now.minute >= 0)
    if not market_closed:
        print(f"ℹ️  当前 {now.strftime('%H:%M')}，A股未收盘（15:00），跳过净值记录")
        return False

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["净值历史"]
    last_row = ws.max_row + 1
    today = now.strftime("%Y-%m-%d")
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] == today:
            print(f"ℹ️  今日已记录，跳过")
            wb.close()
            return False

    ws.cell(row=last_row, column=1, value=today)
    ws.cell(row=last_row, column=2, value=detail.get("A股主账户", 0))
    ws.cell(row=last_row, column=3, value=detail.get("黄金", 0))
    ws.cell(row=last_row, column=4, value=detail.get("账户二（短线）", 0))
    ws.cell(row=last_row, column=5, value=total)
    ws.cell(row=last_row, column=6, value=0)
    ws.cell(row=last_row, column=7, value=pnl)
    ws.cell(row=last_row, column=8, value="收盘记录")
    wb.save(EXCEL_PATH)
    print(f"📝 已记录净值快照: {today}  总资产 {total:,.0f} 元")
    wb.close()
    return True


def load_history():
    """加载净值历史"""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["净值历史"]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        records.append({
            "日期": str(row[0]),
            "A股主账户": row[1] or 0,
            "黄金": row[2] or 0,
            "账户二": row[3] or 0,
            "总资产": row[4] or 0,
            "当日盈亏": row[5] or 0,
            "累计盈亏": row[6] or 0,
        })
    wb.close()
    return records


# ============ 市场数据 ============
# 使用腾讯财经公开 API（无需额外安装依赖）
# 接口: http://qt.gtimg.cn/q=sz000858,sz002850,...
# 返回格式: v_sz000858="...~name~price~...";

STOCK_MARKET_MAP = {
    "6": "sh",  # 上海 600/601/603/605/688
    "0": "sz",  # 深圳 000/001/002
    "3": "sz",  # 创业板 300/301
    "8": "bj",  # 北交所 8
    "4": "bj",  # 北交所 4
}

def _code_to_tdx(code):
    """将6位代码转为腾讯API格式（带市场前缀）"""
    prefix = STOCK_MARKET_MAP.get(code[0], "sz")
    return f"{prefix}{code}"

def fetch_prices():
    """通过腾讯财经 API 获取实时股票价格"""
    import urllib.request
    import urllib.parse
    import json

    prices = {}
    failed = []

    # 股票价格
    stock_names = {v: k for k, v in STOCK_CODES.items()}
    if stock_names:
        tdx_codes = [_code_to_tdx(c) for c in stock_names.keys()]
        url = "http://qt.gtimg.cn/q=" + ",".join(tdx_codes)
        try:
            req = urllib.request.Request(url, headers={
                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"
            })
            resp = urllib.request.urlopen(req, timeout=10)
            text = resp.read().decode("gbk")
            for line in text.strip().split(";"):
                line = line.strip()
                if not line or "=" not in line:
                    continue
                parts = line.split("~")
                if len(parts) >= 4:
                    code_full = parts[0].split("=")[0].strip("v_")
                    code = code_full[2:]  # 去掉 sz/sh 前缀
                    if code in stock_names:
                        try:
                            price = float(parts[3]) if parts[3] else 0
                            if price > 0:
                                prices[stock_names[code]] = price
                            else:
                                failed.append(stock_names[code])
                        except ValueError:
                            failed.append(stock_names[code])
        except Exception as e:
            print(f"⚠️  A股行情获取失败: {e}")
            failed.extend(stock_names.values())

    # 黄金价格（通过 gold-api.com 获取国际金价，换算为人民币/克）
    try:
        req = urllib.request.Request(
            "https://api.gold-api.com/price/XAU/CNY",
            headers={"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"}
        )
        resp = urllib.request.urlopen(req, timeout=10)
        data = json.loads(resp.read().decode("utf-8"))
        price_per_oz = float(data["price"])
        price_per_gram = round(price_per_oz / 31.1035, 2)  # 1金衡盎司 = 31.1035克
        if price_per_gram > 0:
            prices["黄金"] = price_per_gram
        else:
            failed.append("黄金")
    except Exception as e:
        print(f"⚠️ 黄金价格获取失败: {e}")
        failed.append("黄金")

    return prices, failed


def fetch_kline():
    """从新浪财经获取所有持仓股票的历史日K线数据"""
    import urllib.request
    holdings = load_portfolio()
    kline_data = {}
    for h in holdings:
        if h["类别"] != "股票" or not h.get("代码"):
            continue
        code = h["代码"]
        prefix = "sz" if code.startswith(("00", "30")) else "sh"
        url = f"http://money.finance.sina.com.cn/quotes_service/api/json_v2.php/CN_MarketData.getKLineData?symbol={prefix}{code}&scale=240&datalen=45"
        try:
            req = urllib.request.Request(url, headers={
                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"
            })
            resp = urllib.request.urlopen(req, timeout=10)
            raw = resp.read().decode("gbk", errors="ignore")
            if raw:
                import json
                records = json.loads(raw)
                closes = []
                for r in records[-30:]:
                    closes.append({
                        "date": r["day"][:10],
                        "open": float(r["open"]),
                        "high": float(r["high"]),
                        "low": float(r["low"]),
                        "close": float(r["close"]),
                        "volume": int(r.get("volume", 0)),
                    })
                kline_data[h["名称"]] = closes
        except Exception as e:
            print(f"⚠️ {h['名称']} K线获取失败: {e}")
    return kline_data


# ============ 计算 ============
def calculate(holdings):
    """计算各项指标"""
    total_market_value = 0
    total_pnl = 0
    total_cost = 0

    for h in holdings:
        if h["类别"] in ("股票",):
            mv = h["市值"]
            pnl = h["盈亏"]
            total_market_value += mv
            total_pnl += pnl
        elif h["类别"] == "现金":
            total_market_value += h["市值"]
        elif h["类别"] == "黄金":
            mv = h["市值"]
            pnl = h["盈亏"]
            total_market_value += mv
            total_pnl += pnl

    # 按账户汇总
    accounts = {}
    for h in holdings:
        acct = h["账户"]
        if acct not in accounts:
            accounts[acct] = {"市值": 0, "盈亏": 0}
        accounts[acct]["市值"] += h["市值"]
        accounts[acct]["盈亏"] += h["盈亏"]

    return total_market_value, total_pnl, accounts


def update_holdings_with_prices(holdings, prices):
    """用拉取到的价格更新持仓数据"""
    for h in holdings:
        name = h["名称"]
        if name in prices:
            h["现价"] = prices[name]

            # 计算股票市值和盈亏
            if h["类别"] == "股票" and h["数量"] > 0 and h["成本价"] > 0:
                h["市值"] = round(h["数量"] * h["现价"], 2)
                h["盈亏"] = round(h["市值"] - h["数量"] * h["成本价"], 2)
                cost_total = h["数量"] * h["成本价"]
                if cost_total > 0:
                    h["收益率"] = round(h["盈亏"] / cost_total, 4)
                else:
                    h["收益率"] = 0

            # 黄金按克数更新
            elif name == "黄金":
                price = h["现价"] or 0
                if price > 0:
                    if not h["数量"] and h["市值"] and h["市值"] > 0:
                        # 首次获取金价，从当前市值反推持有克数
                        h["数量"] = round(h["市值"] / price, 2)
                        cost_total = h["市值"] - (h["盈亏"] or 0)
                        h["成本价"] = round(cost_total / h["数量"], 2) if h["数量"] > 0 else 0
                    if (h["数量"] or 0) > 0:
                        h["市值"] = round(h["数量"] * price, 2)
                        if h["成本价"] and h["成本价"] > 0:
                            cost_total = round(h["数量"] * h["成本价"], 2)
                            h["盈亏"] = round(h["市值"] - cost_total, 2)
                            h["收益率"] = round(h["盈亏"] / cost_total, 4) if cost_total > 0 else 0

    # 重新计算仓位占比
    total_stock_value = sum(
        h["市值"] for h in holdings
        if h["类别"] in ("股票",) and h["账户"] != "账户二（短线）"
    )
    for h in holdings:
        if h["类别"] in ("股票",) and h["账户"] != "账户二（短线）" and total_stock_value > 0:
            h["仓位占比"] = round(h["市值"] / total_stock_value, 4)
        elif h["类别"] == "现金":
            h["仓位占比"] = 0

    return holdings


# ============ 输出 ============
def print_holdings(holdings):
    """打印持仓表格"""
    total_value, total_pnl, accounts = calculate(holdings)

    print("\n" + "=" * 80)
    print(f"📊 投资持仓全景  ({datetime.now().strftime('%Y-%m-%d %H:%M')})")
    print("=" * 80)

    print(f"\n{'资产名称':<12} {'现价':>10} {'成本':>10} {'数量':>8} {'市值':>12} {'盈亏':>12} {'收益率':>10} {'占比':>8}")
    print("-" * 82)
    for h in holdings:
        if h["类别"] not in ("股票",):
            continue
        name = h["名称"]
        price = f"{h['现价']:.2f}" if h["现价"] else "-"
        cost = f"{h['成本价']:.2f}" if h["成本价"] and h["成本价"] != 0 else "-"
        qty = f"{h['数量']:.0f}" if h["数量"] and h["数量"] != 0 else "-"
        mv = f"{h['市值']:>10,.0f}" if h["市值"] else "-"
        pnl_val = h["盈亏"] if h["盈亏"] else 0
        pnl_str = f"{pnl_val:>+10,.0f}" if pnl_val != 0 else f"{pnl_val:>10,.0f}"
        ret = f"{h['收益率']*100:>+7.1f}%" if h["收益率"] and h["收益率"] != 0 else "    -"
        alloc = f"{h['仓位占比']*100:>5.1f}%" if h["仓位占比"] and h["仓位占比"] != 0 else "   -"
        print(f"{name:<12} {price:>10} {cost:>10} {qty:>8} {mv:>12} {pnl_str:>12} {ret:>10} {alloc:>8}")

    print("-" * 82)

    # 汇总
    print(f"\n股票市值: {total_value:>10,.0f} 元    浮亏: {total_pnl:>+10,.0f} 元")

    print(f"\n📂 按账户:")
    for acct, data in accounts.items():
        print(f"  {acct:<14}  资产 {data['市值']:>8,.0f} 元    {'盈亏' if data['盈亏'] >= 0 else '浮亏'} {data['盈亏']:>+8,.0f} 元")

    print(f"\n💰 总资产: {sum(a['市值'] for a in accounts.values()):>8,.0f} 元")
    print(f"📉 累计盈亏: {sum(a['盈亏'] for a in accounts.values()):>+8,.0f} 元")


def print_allocation(holdings):
    """打印资产配置"""
    print(f"\n📊 资产配置分析 ({datetime.now().strftime('%Y-%m-%d %H:%M')})")
    print("=" * 60)

    by_type = {}
    total = 0
    for h in holdings:
        t = h["类别"]
        if t not in by_type:
            by_type[t] = 0
        by_type[t] += h["市值"]
        total += h["市值"]

    for t, v in sorted(by_type.items(), key=lambda x: -x[1]):
        pct = v / total * 100 if total > 0 else 0
        bar = "█" * int(pct / 5) + "░" * (20 - int(pct / 5))
        print(f"  {t:<8} {bar} {v:>8,.0f} 元 ({pct:>5.1f}%)")

    print(f"\n  总资产: {total:>8,.0f} 元")


def print_history(records):
    """打印净值历史"""
    if not records:
        print("\n暂无净值历史记录。运行 `python3 portfolio_agent.py snapshot` 记录第一条。")
        return
    print(f"\n📈 净值趋势")
    print("=" * 80)
    print(f"{'日期':<14} {'总资产':>12} {'A股主账户':>12} {'黄金':>12} {'短线':>12} {'累计盈亏':>12}")
    print("-" * 80)
    for r in records:
        print(f"{r['日期']:<14} {r['总资产']:>12,.0f} {r['A股主账户']:>12,.0f} {r['黄金']:>12,.0f} {r['账户二']:>12,.0f} {r['累计盈亏']:>12,+.0f}")


def print_report(holdings):
    """生成完整报告"""
    total_value, total_pnl, accounts = calculate(holdings)
    by_type = {}
    for h in holdings:
        t = h["类别"]
        by_type.setdefault(t, {"市值": 0, "盈亏": 0})
        by_type[t]["市值"] += h["市值"]
        by_type[t]["盈亏"] += h.get("盈亏", 0)

    print("\n" + "=" * 60)
    print(f"📋 投资组合报告")
    print(f"   生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 60)

    print(f"\n📊 总览")
    print(f"  总资产:       {total_value:>10,.0f} 元")
    print(f"  累计盈亏:     {total_pnl:>+10,.0f} 元")
    print(f"  总收益率:     {total_pnl/(total_value-total_pnl)*100:>+7.2f}%" if total_value != total_pnl else "  总收益率:      -")

    print(f"\n📂 各账户情况")
    for acct, data in sorted(accounts.items()):
        pnl = data["盈亏"]
        cost_base = data["市值"] - pnl
        ret = pnl / cost_base * 100 if cost_base > 0 else 0
        status = "🔴" if pnl < 0 else "🟢"
        print(f"  {status} {acct:<14}  {data['市值']:>8,.0f} 元    {'亏损' if pnl < 0 else '盈利'}: {pnl:>+8,.0f} 元 ({ret:>+.1f}%)")

    print(f"\n📊 资产配置")
    for t, d in sorted(by_type.items(), key=lambda x: -x[1]["市值"]):
        pct = d["市值"] / total_value * 100
        bar = "▓" * int(pct / 5) + "░" * (20 - int(pct / 5))
        print(f"  {t:<8} {bar} {d['市值']:>8,.0f} 元 ({pct:>5.1f}%)")

    # 风险提示
    if total_pnl < 0:
        print(f"\n⚠️  当前整体浮亏 {abs(total_pnl):,.0f} 元")

    # 集中度提示
    for h in holdings:
        if h["仓位占比"] and h["仓位占比"] > 0.4:
            print(f"⚠️  {h['名称']} 仓位占比 {h['仓位占比']*100:.1f}%，超过 40%，集中度偏高")


# ============ 风险诊断 ============
def risk_analysis(holdings):
    """诊断持仓风险，返回风险列表"""
    risks = []
    total_value, total_pnl, accounts = calculate(holdings)

    # 1. 单票集中度
    main_stock_value = sum(
        h["市值"] for h in holdings
        if h["类别"] == "股票" and h["账户"] != "账户二（短线）"
    )
    for h in holdings:
        if h["类别"] == "股票" and h["仓位占比"] and h["仓位占比"] > 0.25:
            price = h["现价"] or 0
            qty = h["数量"] or 0
            target_mv = main_stock_value * 0.20
            excess_mv = h["市值"] - target_mv
            if price > 0 and qty > 0 and excess_mv > 0:
                sell_shares = int(excess_mv / price / 100) * 100     # 向下取整到整手
                if sell_shares <= 0:
                    sell_shares = 100                                 # 至少卖 1 手
                if sell_shares >= qty:
                    sell_shares = qty
                sell_amount = sell_shares * price
                remain = qty - sell_shares
                remain_amount = remain * price
                new_pct = remain_amount / main_stock_value * 100
                suggestion = (
                    f"建议减仓 {sell_shares} 股（{sell_shares//100}手），"
                    f"当前 {qty} 股 → 保留 {remain} 股，"
                    f"释放约 {sell_amount:,.0f} 元，仓位从 {h['仓位占比']*100:.1f}% 降至约 {new_pct:.1f}%"
                )
            else:
                suggestion = "建议分批减仓至 15-25%"
            risks.append({
                "级别": "🔴 高风险",
                "问题": f"{h['名称']} 仓位占比 {h['仓位占比']*100:.1f}%，超过建议上限 25%",
                "建议": suggestion,
            })

    # 2. 行业集中度（新能源车产业链）
    new_energy_stocks = ["科达利", "三花智控", "拓普集团"]
    ne_value = sum(
        h["市值"] for h in holdings
        if h["名称"] in new_energy_stocks and h["类别"] == "股票"
    )
    total_stock_value = sum(h["市值"] for h in holdings if h["类别"] == "股票" and h["账户"] != "账户二（短线）")
    if total_stock_value > 0 and ne_value / total_stock_value > 0.5:
        risks.append({
            "级别": "🔴 高风险",
            "问题": f"新能源车产业链（科达利+三花+拓普）占比 {ne_value/total_stock_value*100:.1f}%，超过 50%",
            "建议": f"建议新增消费/医药/科技等行业股票，将同行业占比降到 40% 以下",
        })

    # 3. 深度亏损检查（浮亏 > 30%）
    for h in holdings:
        if h["类别"] == "股票" and h["收益率"] and h["收益率"] < -0.3:
            price = h["现价"] or 0
            qty = h["数量"] or 0
            if price > 0 and qty > 0:
                first_target = price * 1.10
                first_sell = 100 if qty >= 100 else qty
                second_sell = (qty - first_sell) if qty - first_sell >= 100 else 0
                if second_sell > 0:
                    second_target = price * 1.20
                    suggestion = (
                        f"反弹减仓计划（当前 {qty} 股）:\n"
                        f"  第一档：涨至 {first_target:.1f} 元（+10%），卖 {first_sell} 股（{first_sell//100}手）\n"
                        f"  第二档：涨至 {second_target:.1f} 元（+20%），卖 {second_sell} 股（{second_sell//100}手）\n"
                        f"  建议设好条件单，不要被动持有等回本"
                    )
                else:
                    suggestion = (
                        f"反弹止损计划（当前 {qty} 股）:\n"
                        f"  反弹至 {first_target:.1f} 元（+10%）时全部卖出 {qty} 股"
                    )
            else:
                suggestion = "设置反弹减仓计划：每涨 10-15% 减 1/3"
            risks.append({
                "级别": "🟡 中风险",
                "问题": f"{h['名称']} 浮亏 {h['收益率']*100:.1f}%（亏损 {abs(h['盈亏']):,.0f} 元），深度套牢",
                "建议": suggestion,
            })

    # 4. 现金比例
    cash_value = sum(h["市值"] for h in holdings if h["类别"] == "现金")
    if total_value > 0 and cash_value / total_value < 0.08:
        risks.append({
            "级别": "🟡 中风险",
            "问题": f"现金仅 {cash_value:,.0f} 元（占比 {cash_value/total_value*100:.1f}%），低于建议的 10%",
            "建议": "建议至少保留 2-3 万元现金作为应急储备",
        })

    # 5. 无固收类资产
    has_bond = any(h["类别"] in ("债基", "固收", "理财") for h in holdings)
    if not has_bond:
        risks.append({
            "级别": "🟡 中风险",
            "问题": "缺少固收/债券类资产，组合波动率偏高",
            "建议": "建议配置 15-20% 短债基金或银行理财作为压舱石",
        })

    # 6. 短线账户亏损率
    for acct, data in sorted(accounts.items()):
        if "短线" in acct:
            cost = data["市值"] - data["盈亏"]
            if cost > 0 and data["盈亏"] / cost < -0.2:
                risks.append({
                    "级别": "🔴 高风险",
                    "问题": f"{acct} 亏损率达 {data['盈亏']/cost*100:.1f}%（亏损 {abs(data['盈亏']):,.0f} 元）",
                    "建议": "建议暂停短线操作，剩余资金并入主账户做中线持有",
                })

    return risks


def breakeven_analysis(holdings, cash=0):
    """分析每只股票的盈亏进度和预估回本时间（交易日天数），加仓分析"""
    results = []
    for h in holdings:
        if h["类别"] != "股票" or not h.get("数量"):
            continue
        cost = h["成本价"]
        price = h["现价"]
        qty = h["数量"]
        if not cost or not price or cost <= 0 or price <= 0 or not qty:
            continue

        total_cost = round(cost * qty, 2)
        total_value = round(price * qty, 2)
        pnl = round(total_value - total_cost, 2)
        return_pct = (total_cost - total_value) / total_cost * 100

        # 回本进度 = 现价/成本价
        be_progress = round(price / cost * 100, 1)

        # 还需涨多少才能回本
        rise_needed = (cost / price - 1) * 100 if price > 0 else 0

        # 按日涨幅预估回本时间（交易日，月均约22天）
        scenarios = {}
        for label, desc, daily_rise in [
            ("optimistic", "乐观 +5%/月", 5 / 22),
            ("moderate", "中性 +3%/月", 3 / 22),
            ("conservative", "保守 +1%/月", 1 / 22),
        ]:
            if daily_rise > 0 and rise_needed > 0:
                days = round(rise_needed / daily_rise)
                scenarios[label] = {"label": desc, "days": days}

        results.append({
            "name": h["名称"],
            "account": h["账户"],
            "cost": cost,
            "price": price,
            "qty": qty,
            "total_cost": total_cost,
            "total_value": total_value,
            "pnl": pnl,
            "return_pct": round(return_pct, 2),
            "be_progress": be_progress,
            "rise_needed": round(rise_needed, 1),
            "scenarios": scenarios,
            "recommendation": {"cash_available": round(cash, 2)},
        })
    return results


def generate_advice(holdings, risks):
    """生成短/中/长期理财建议"""
    total_value, total_pnl, accounts = calculate(holdings)
    by_type = {}
    for h in holdings:
        t = h["类别"]
        by_type[t] = by_type.get(t, 0) + h["市值"]

    now = datetime.now().strftime("%Y-%m-%d")
    advice = f"""
╔══════════════════════════════════════════════════════════════╗
║             📋 每日投资建议 · {now}              ║
╚══════════════════════════════════════════════════════════════╝

"""

    # ====== 短期建议（1-4周） ======
    advice += "━" * 56 + "\n"
    advice += "  🔹 短期建议（未来 1-4 周）\n"
    advice += "━" * 56 + "\n\n"

    # 仓位警告
    advice += "  ▌仓位纪律\n"
    has_warning = False
    for r in risks:
        if "仓位占比" in r["问题"]:
            advice += f"    {r['级别']} {r['问题']}\n"
            advice += f"    → {r['建议']}\n\n"
            has_warning = True
    if not has_warning:
        advice += "    ✅ 单票仓位均在 25% 以内，良好\n\n"

    # 深度亏损处理
    advice += "  ▌深度持仓处理\n"
    has_deep_loss = False
    for h in holdings:
        if h["类别"] == "股票" and h["收益率"] and h["收益率"] < -0.3:
            price = h["现价"] or 0
            qty = h["数量"] or 0
            advice += f"    🟡 {h['名称']}：现价 {price:.2f}，浮亏 {h['收益率']*100:.1f}%（{qty} 股）\n"
            if price > 0 and qty > 0:
                first_target = price * 1.10
                first_sell = 100 if qty >= 100 else qty
                second_sell = qty - first_sell if qty - first_sell >= 100 else 0
                if second_sell > 0:
                    advice += (
                        f"       → 反弹减仓计划（A股最小 1手=100股）:\n"
                        f"          第一档：涨至 {first_target:.1f} 元（+10%），卖 {first_sell} 股（{first_sell//100}手）\n"
                        f"          第二档：涨至 {price*1.20:.1f} 元（+20%），卖 {second_sell} 股（{second_sell//100}手）\n"
                        f"          不要被动持有等回本，设好条件单\n\n"
                    )
                else:
                    advice += (
                        f"       → 反弹止损：涨至 {first_target:.1f} 元（+10%）时全部卖出 {qty} 股\n\n"
                    )
            else:
                advice += f"       → 建议设反弹减仓线：{h['现价']*1.1:.1f}（+10%）、{h['现价']*1.2:.1f}（+20%）\n\n"
            has_deep_loss = True
    if not has_deep_loss:
        advice += "    ✅ 无深度套牢持仓\n\n"

    # 短线
    for acct, data in sorted(accounts.items()):
        if "短线" in acct:
            advice += "  ▌短线账户\n"
            cost = data["市值"] - data["盈亏"]
            advice += f"    🟡 {acct}：资产 {data['市值']:,.0f} 元，累计亏损 {abs(data['盈亏']):,.0f} 元\n"
            if cost > 0 and data["盈亏"] / cost < -0.2:
                advice += f"       → 建议：暂停短线交易，剩余资金并入主账户\n\n"

    # 近期操作建议
    advice += "  ▌近期操作\n"
    advice += "    • 每周检查一次持仓，单票超过 25% 触发减仓信号\n"
    advice += "    • 注意 A 股最小交易单位 1手 = 100股，减仓时按整手操作\n"
    advice += "    • 五粮液、天齐锂业设反弹减仓计划，不要被动等待回本\n"
    advice += "    • 保留至少 2 万现金，不要满仓\n\n"

    # ====== 中期建议（1-6个月） ======
    advice += "━" * 56 + "\n"
    advice += "  🔸 中期建议（未来 1-6 个月）\n"
    advice += "━" * 56 + "\n\n"

    # 行业配置
    stock_names = [h["名称"] for h in holdings if h["类别"] == "股票" and h["账户"] != "账户二（短线）"]
    ne_count = sum(1 for s in stock_names if s in ["科达利", "三花智控", "拓普集团"])
    advice += "  ▌行业分散\n"
    if ne_count >= 3:
        advice += "    🔴 新能源车产业链持仓过重（科达利+三花+拓普）\n"
        advice += "    → 中期目标：减仓 1-2 只新能源票，换入以下行业：\n"
        advice += "      • 大消费：白酒（五粮液已有）、家电、食品饮料\n"
        advice += "      • 医药：医疗器械、创新药（防御属性）\n"
        advice += "      • 科技/互联网：立讯已有，可考虑加仓或换ETF\n\n"

    # 板块分析
    advice += "  ▌板块分析（当前持仓结构）\n"
    from collections import defaultdict
    sector_groups = defaultdict(list)
    for s in holdings:
        if s["类别"] == "股票" and s["账户"] != "账户二（短线）" and s["名称"] in SECTOR_MAP:
            sector_groups[SECTOR_MAP[s["名称"]]].append(s)
    sector_total = {s: sum(x["市值"] for x in stocks) for s, stocks in sector_groups.items()}
    total_mv = sum(sector_total.values())
    for sector, total in sorted(sector_total.items(), key=lambda x: -x[1]):
        stocks = sector_groups[sector]
        names_str = ", ".join(f"{s['名称']}{s['市值']:,.0f}元" for s in stocks)
        pct = total / total_mv * 100 if total_mv > 0 else 0
        advice += f"    {sector}\n"
        advice += f"      市值 {total:,.0f} 元，占股票仓位 {pct:.1f}%\n"
        outlook = SECTOR_OUTLOOK.get(sector, "")
        if outlook:
            advice += f"      预期：{outlook}\n"
        advice += "\n"
    advice += "  ▌资产再平衡\n"
    advice += f"    当前: 股票 {by_type.get('股票',0)/total_value*100:.0f}% | 黄金 {by_type.get('黄金',0)/total_value*100:.0f}% | 现金 {by_type.get('现金',0)/total_value*100:.0f}%\n"
    advice += "    目标: 股票 55-60% | 黄金 10-15% | 固收 15-20% | 现金 10%\n"
    advice += "    → 中期应逐步加入债基/理财作为压舱石\n\n"

    # 黄金
    gold_value = by_type.get("黄金", 0)
    gold_pct = gold_value / total_value * 100 if total_value > 0 else 0
    if gold_pct > 18:
        advice += "  ▌黄金持仓\n"
        advice += f"    🟡 黄金占比 {gold_pct:.1f}%，高于建议的 10-15%\n"
        advice += "    → 黄金更多是对冲工具，建议维持 10-15% 即可\n"
        advice += "    → 如金价短期大涨可适当减持锁定利润\n\n"

    # ====== 长期建议（6个月以上） ======
    advice += "━" * 56 + "\n"
    advice += "  🔹 长期建议（6 个月以上）\n"
    advice += "━" * 56 + "\n\n"

    advice += "  ▌资产配置框架\n\n"
    advice += "    按 24 万总资产，建议终极目标：\n\n"
    advice += "    ┌────────────────────────────────────────────┐\n"
    advice += "    │  股票（8-10只分散）      55-60%  13-14万  │\n"
    advice += "    │  ├─ 单票上限 20-25%                      │\n"
    advice += "    │  ├─ 不同行业 4-5 个                     │\n"
    advice += "    │  └─ 避免超过 3 只同产业链               │\n"
    advice += "    ├────────────────────────────────────────────┤\n"
    advice += "    │  黄金                     10-15%   2.5-3.5万│\n"
    advice += "    ├────────────────────────────────────────────┤\n"
    advice += "    │  债基/银行理财            15-20%   3.5-5万 │\n"
    advice += "    ├────────────────────────────────────────────┤\n"
    advice += "    │  现金                     10%     2-2.5万  │\n"
    advice += "    └────────────────────────────────────────────┘\n\n"

    advice += "  ▌风险底线\n"
    advice += "    • 永远不要单票超过 30%\n"
    advice += "    • 永远不要满仓（至少留 10% 现金）\n"
    advice += "    • 永远不要用生活费、急用钱投资\n"
    advice += "    • 如果一笔投资让你睡不着觉，说明仓位太重\n\n"

    advice += "  ▌长期纪律\n"
    advice += "    • 每月定投：建议每月收入的 10-20% 定投宽基指数（沪深300/中证500）\n"
    advice += "    • 季度再平衡：每季度检查一次资产比例，偏离超过 5% 就调整\n"
    advice += "    • 年度复盘：每年底做一次全面财务复盘\n\n"

    # 风险摘要
    advice += "━" * 56 + "\n"
    advice += "  ⚠️  待处理风险（{0} 项）\n".format(len(risks))
    advice += "━" * 56 + "\n\n"
    for r in risks:
        advice += f"  {r['级别']} {r['问题']}\n"
        advice += f"    → {r['建议']}\n\n"

    advice += "━" * 56 + "\n"
    advice += "  💡 自动生成，仅供参考。投资有风险，决策需谨慎。\n"
    advice += "━" * 56 + "\n"

    return advice


def save_advice(advice_text):
    """保存每日建议到文件"""
    today = datetime.now().strftime("%Y-%m-%d")
    output_dir = DATA_DIR / "每日建议"
    output_dir.mkdir(exist_ok=True)
    path = output_dir / f"{today}.txt"
    path.write_text(advice_text, encoding="utf-8")
    print(f"📝 建议已保存: {path}")
    return path


def print_daily(holdings):
    """每日简报：先刷新行情 → 记录净值 → 风险诊断 → 生成建议"""
    print("=" * 60)
    print(f"  📋 投资追踪 · 每日简报")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 60)

    # 1. 行情快照
    print("\n🔄 正在获取实时行情...")
    prices, failed = fetch_prices()
    if prices:
        holdings = update_holdings_with_prices(holdings, prices)
        save_prices(holdings)
        print(f"✅ 已更新 {len(prices)} 个价格")
    if failed:
        print(f"⚠️  以下未获取到: {', '.join(failed)}")

    # 2. 持仓概览
    print_holdings(holdings)

    # 3. 记录净值
    total_value, total_pnl, accounts = calculate(holdings)
    detail = {k: v["市值"] for k, v in accounts.items()}
    save_networth_snapshot(total_value, total_pnl, detail)

    # 4. 风险诊断
    risks = risk_analysis(holdings)
    print(f"\n{'='*60}")
    print(f"  ⚠️  风险诊断（发现 {len(risks)} 项）")
    print(f"{'='*60}")
    for r in risks:
        print(f"\n  {r['级别']}")
        print(f"  问题: {r['问题']}")
        print(f"  建议: {r['建议']}")

    # 5. 生成建议
    advice = generate_advice(holdings, risks)
    print(advice)
    save_advice(advice)

    # 6. 同步更新 HTML 看板
    try:
        from report_generator import generate as gen_html
        gen_html()
        print(f"🌐 HTML 看板已更新")
    except Exception:
        pass
def main():
    if len(sys.argv) < 2:
        print("用法:")
        print("  python3 portfolio_agent.py holdings    # 查看持仓")
        print("  python3 portfolio_agent.py refresh     # 更新实时价格")
        print("  python3 portfolio_agent.py report      # 完整报告")
        print("  python3 portfolio_agent.py allocation  # 资产配置")
        print("  python3 portfolio_agent.py history     # 净值历史")
        print("  python3 portfolio_agent.py snapshot    # 保存净值快照")
        print("  python3 portfolio_agent.py risks       # 风险诊断")
        print("  python3 portfolio_agent.py advise      # 理财建议")
        print("  python3 portfolio_agent.py daily       # 每日简报（行情+建议+记录）")
        return

    cmd = sys.argv[1]

    if cmd == "holdings":
        holdings = load_portfolio()
        print_holdings(holdings)

    elif cmd == "refresh":
        print("🔄 正在获取实时行情...")
        holdings = load_portfolio()
        prices, failed = fetch_prices()
        if prices:
            print(f"✅ 获取到 {len(prices)} 个价格")
            holdings = update_holdings_with_prices(holdings, prices)
            save_prices(holdings)
            print_holdings(holdings)
        if failed:
            print(f"⚠️  以下资产未获取到价格: {', '.join(failed)}")

    elif cmd == "report":
        holdings = load_portfolio()
        print_report(holdings)

    elif cmd == "allocation":
        holdings = load_portfolio()
        print_allocation(holdings)

    elif cmd == "history":
        records = load_history()
        print_history(records)

    elif cmd == "snapshot":
        holdings = load_portfolio()
        total_value, total_pnl, accounts = calculate(holdings)
        detail = {k: v["市值"] for k, v in accounts.items()}
        save_networth_snapshot(total_value, total_pnl, detail)
        print(f"  总资产: {total_value:>8,.0f} 元  累计盈亏: {total_pnl:>+8,.0f} 元")

    elif cmd == "risks":
        holdings = load_portfolio()
        risks = risk_analysis(holdings)
        print(f"\n{'='*60}")
        print(f"  ⚠️  风险诊断（发现 {len(risks)} 项）")
        print(f"{'='*60}")
        for r in risks:
            print(f"\n  {r['级别']}")
            print(f"  问题: {r['问题']}")
            print(f"  建议: {r['建议']}")
        if not risks:
            print("  ✅ 无显著风险，继续保持")

    elif cmd == "advise":
        holdings = load_portfolio()
        risks = risk_analysis(holdings)
        advice = generate_advice(holdings, risks)
        print(advice)
        save_advice(advice)

    elif cmd == "daily":
        holdings = load_portfolio()
        print_daily(holdings)

    else:
        print(f"未知命令: {cmd}")


if __name__ == "__main__":
    main()
